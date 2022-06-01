# -*- coding: utf8 -*-
__author__ = 'wangqiang'


"""
对划词结果输出文件中结果类整理

使用前提：
    安装python3 环境
    安装python包openpyxl
    
使用示例：
python3 remake_cut_word_export.py export_result.xlsx 3 4
    参数说明：
        export_result.xlsx 需要重构的文件
        3 划词结果所在列（默认3）
        4 结果重构后保存列（默认4）
"""

import sys
import json
from openpyxl import load_workbook, workbook


def result_to_json(txt):
    """
    将划词结果转换成json串
    :param txt:
    :return:
    """
    entities = []
    index, real_index = 0, 0
    left_found, middle_found, right_found = -1, -1, -1

    while True:
        if index >= len(txt):
            break

        if txt[index] == "[" and left_found == -1:
            # 标记划词开始
            left_found = index

        if txt[index] == "]" and left_found > -1:
            # 标记中间分隔
            middle_found = index

        if txt[index] == " ":
            if middle_found > 0:
                right_found = index
            else:
                # 当划词未开始时，空格不作为有效字符计算，直接跳过
                index += 1
                continue

        if left_found == -1:
            real_index += 1
        else:
            if -1 < left_found < middle_found < right_found:
                # 找到一个划词记录
                entity = txt[left_found + 1: middle_found]
                etype = txt[middle_found + 2: right_found]
                eindex = real_index
                entities.append({
                    "entity": entity,
                    "type": etype,
                    "index": (eindex, eindex + len(entity) - 1)
                })
                real_index += len(entity)
                left_found, middle_found, right_found = -1, -1, -1
        index += 1
    return json.dumps(entities, ensure_ascii=False)


if __name__ == '__main__':
    export_excel = "1.xlsx"
    if len(sys.argv) > 1:
        export_excel = sys.argv[1]
    remake_column = 3
    if len(sys.argv) > 2:
        remake_column = int(sys.argv[2])
    result_column = 4
    if len(sys.argv) > 3:
        result_column = int(sys.argv[3])
    if result_column <= remake_column:
        result_column = remake_column + 1

    swb = load_workbook(export_excel)

    twb = workbook.Workbook()

    sheetnames = swb.sheetnames

    for sname in sheetnames:
        ssheet = swb[sname]
        max_row = ssheet.max_row
        max_column = ssheet.max_column

        tsheet = twb.active
        tsheet.title = sname

        for row_index in range(1, max_row + 1):
            for column_index in range(1, max_column + 1):
                val = ssheet.cell(row=row_index, column=column_index).value
                tsheet.cell(row=row_index, column=column_index).value = val
                # 需要重构的列
                if column_index == remake_column:
                    json_txt = result_to_json(val)
                    print(f"find entities:{json_txt}")
                    tsheet.cell(row=row_index, column=result_column).value = json_txt

    remake_file = f"remake_{export_excel}"
    twb.save(remake_file)
    print("mission completed.")
