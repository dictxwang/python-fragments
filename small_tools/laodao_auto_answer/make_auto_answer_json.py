#!/usr/bin/env python3
# -*- coding: utf8 -*-
__author__ = 'wq'

from openpyxl import load_workbook
import json


'''
生成答题js所需的json串（基于"八抓20项安全考试.xlsx"）
实际就是题目和答案的对应关系
'''


def parse_single_choice(sheet):
    '''
    单选
    :return:
    '''
    max_row = sheet.max_row
    max_column = sheet.max_column
    qa = {}
    # 正确答案在第5列
    for row in range(2, max_row + 1):
        question = sheet.cell(row=row, column=1).value
        question = question.strip()
        answer_flag = sheet.cell(row=row, column=5).value
        for column in range(6, max_column + 1):
            text = sheet.cell(row=row, column=column).value
            if text and text[0:1] == answer_flag:
                answer = text[2:].strip()
                qa[question] = [answer]
                break
    return qa


def parse_multi_choice(sheet):
    '''
    多选
    :return:
    '''
    max_row = sheet.max_row
    max_column = sheet.max_column
    qa = {}
    # 正确答案在第5列
    for row in range(2, max_row + 1):
        question = sheet.cell(row=row, column=1).value
        question = question.strip()
        answer_flag = sheet.cell(row=row, column=5).value
        answer_flag_lst = list(answer_flag)
        answer_lst = []
        for column in range(6, max_column + 1):
            text = sheet.cell(row=row, column=column).value
            if text and text[0:1] in answer_flag_lst:
                answer = text[2:].strip()
                answer_lst.append(answer)
        qa[question] = answer_lst
    return qa


def parse_judge(sheet):
    '''
    判断
    :return:
    '''
    max_row = sheet.max_row
    qa = {}
    # 正确答案在第5列
    for row in range(2, max_row + 1):
        question = sheet.cell(row=row, column=1).value
        question = question.strip()
        answer = sheet.cell(row=row, column=5).value
        answer = answer.strip()
        qa[question] = [answer]
    return qa


if __name__ == '__main__':
    efile = "八抓20项安全考试.xlsx"
    wb = load_workbook(efile)
    sheet_names = wb.sheetnames
    supports_types = ["单选题", "多选题", "判断题"]

    question_answer = {}
    for sname in sheet_names:
        if sname not in supports_types:
            print(f"not support {sname}")
            continue
        sheet = wb[sname]
        qa = None
        if sname == "单选题":
            qa = parse_single_choice(sheet)
        elif sname == "多选题":
            qa = parse_multi_choice(sheet)
        elif sname == "判断题":
            qa = parse_judge(sheet)
        for k, v in qa.items():
            question_answer[k] = v
    print(json.dumps(question_answer, ensure_ascii=False))
