#!/usr/bin/env python3
# -*- coding: utf8 -*-
__author__ = 'wangqiang'

'''
pip install openpyxl==3.0.9
'''

from datetime import datetime
import uuid
from openpyxl import load_workbook, workbook


def read_wb():
    print("\n[openpyxl_read_wb]")

    rfile = "data/openpyxl_sample.xlsx"
    wb = load_workbook(rfile)
    # 获取所有sheet的name，旧版本的库使用 wb.get_sheet_names()
    sheet_names = wb.sheetnames

    # 获取当前活动的sheet（焦点所在的sheet页）
    sheet_active = wb.active
    print("active sheet is {}".format(sheet_active.title))

    for sname in sheet_names:
        print("\nreading from {}".format(sname))
        # 旧版本库使用 wb.get_sheet_by_name("")
        sheet = wb[sname]
        max_row = sheet.max_row
        max_column = sheet.max_column
        # 遍历单元格，行列下标从1开始
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                val = sheet.cell(row=row, column=column).value
                # 也可以通过cell名访问cell
                # sheet["A1"]
                if val:
                    cell = sheet.cell(row=row, column=column)
                    # cell.is_date 判断是否是时间类型
                    print("<{},{}> {} is_date={}".format(row, column, val, cell.is_date))
                else:
                    print("<{},{}> None".format(row, column))
    wb.close()


def append_wb():
    print("\n[openpyxl_append_wb]")

    rfile = "data/openpyxl_sample.xlsx"
    wb = load_workbook(rfile)
    sheet = wb["Sheet2"]
    # 获取当前最大行
    max_row = sheet.max_row
    for i in range(0, 2):
        now = datetime.now()
        now_format = datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S.%f")
        text = str(uuid.uuid4())
        sheet.cell(row=max_row+i, column=1).value = now_format
        # 时间类型会自动转换格式(cell.is_date 属性为True)
        sheet.cell(row=max_row+i, column=2).value = now
        sheet.cell(row=max_row+i, column=3, value=text)

    # 保存到原文件，也可以保存到新的文件
    wb.save(rfile)
    wb.close()


def create_wb():
    print("\n[openpyxl_create_wb]")
    nfile = "data/openpyxl_new.xlsx"

    wb = workbook.Workbook()
    sheet = wb.active

    # 重命名sheet名（新建默认为Sheet）
    sheet.title = "Sheet1"
    sheet["A1"] = "A"
    sheet["b1"] = "B"
    sheet.cell(row=1, column=3).value = "C"

    # 合并单元格，合并范围 A2:D3
    sheet.merge_cells("A2:D3")
    # 合并单元格（如果合并位置出现冲突，将破坏excel，打开时会提示错误）
    sheet.merge_cells("E4:F6")

    # 给合并后区域填充内容
    sheet["A2"] = "Merged"
    sheet["B2"] = "Merged"  # B2已经不存在，填充无效

    # 获取合并的单元格区域（如果是多个合并区域将返回多个值）。旧版本方法sheet.merged_cell_ranges
    merged_ranges = sheet.merged_cells.ranges
    for m in merged_ranges:
        print(m)  # A2:D3\nE4:F6

    wb.save(nfile)


if __name__ == '__main__':

    read_wb()
    append_wb()
    create_wb()
